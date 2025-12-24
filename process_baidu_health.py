import openpyxl
import os

# 定义Excel文件路径
excel_file = "/Users/wangyucheng02/Documents/我的开放项目/IronMan/百度大健康广告.csv"

# 检查文件是否存在
if not os.path.exists(excel_file):
    print(f"文件不存在: {excel_file}")
    exit()

print("--- 开始读取Excel文件 ---\n")

# 打开Excel文件
workbook = openpyxl.load_workbook(excel_file)

# 获取所有工作表名称
print("工作表名称:")
for sheet_name in workbook.sheetnames:
    print(f"  - {sheet_name}")

# 选择第一个工作表
sheet = workbook.active
print(f"\n当前工作表: {sheet.title}")

# 获取工作表的行数和列数
max_row = sheet.max_row
max_col = sheet.max_column
print(f"工作表大小: {max_row} 行, {max_col} 列")

# 读取前5行数据（包括表头）
print(f"\n前5行数据:\n")

# 打印表头
print("表头:")
for col in range(1, max_col + 1):
    cell_value = sheet.cell(row=1, column=col).value
    print(f"  列{col}: {cell_value}")

# 打印前4行数据（不包括表头）
print(f"\n前4行数据:")
for row in range(2, min(6, max_row + 1)):
    print(f"\n第 {row-1} 条数据:")
    for col in range(1, max_col + 1):
        cell_value = sheet.cell(row=row, column=col).value
        header = sheet.cell(row=1, column=col).value
        print(f"  {header}: {cell_value}")

# 关闭Excel文件
workbook.close()

print(f"\n--- 文件读取完成 ---\n")
print(f"注意: 该文件包含 {max_row-1} 条数据（不包括表头）")
